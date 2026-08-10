"""Write the flagged players to a formatted .xlsx.

openpyxl is the one dependency in this project and it is optional: everything
still works as CSV without it. It is imported inside the function so that
`import ootp_scout` never fails on a machine that lacks it.
"""

from __future__ import annotations

from .flagging import Finding, GroupFit

# Thresholds in standard deviations. A player's row is tinted by how far his
# projection sits above what his grade predicts.
STRONG_Z = 2.0
NOTABLE_Z = 1.0

COLUMNS = [
    ("Rank", 6), ("Player", 24), ("Pos", 6), ("Age", 6), ("Group", 10),
    ("Grade", 8), ("Projected WAR", 15), ("Expected WAR", 14),
    ("Differential", 13), ("z", 8), ("Scouting Accuracy", 18),
]

# Added after Projected WAR when the sheet contains pitchers.
PITCHER_COLUMNS = [("rWAR", 9), ("rWAR - WAR", 13)]


class SpreadsheetUnavailable(RuntimeError):
    pass


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as error:  # pragma: no cover - environment dependent
        raise SpreadsheetUnavailable(
            "writing .xlsx needs openpyxl (pip install openpyxl). Use a .csv "
            "filename instead to stay dependency-free.") from error
    return openpyxl


def _write_players_sheet(sheet, findings: list[Finding], grade_label: str,
                         strong_z: float, notable_z: float,
                         rating_columns: list[str]) -> None:
    """Every player in the pool, best differential first, tinted both ways.

    One table rather than two: splitting underrated and overrated onto separate
    sheets hid the middle of the distribution entirely, and the middle is what
    makes the extremes legible.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    strong_up = PatternFill("solid", fgColor="FFC7EFCE")     # green
    notable_up = PatternFill("solid", fgColor="FFFFF2CC")    # amber
    strong_down = PatternFill("solid", fgColor="FFFFC7CE")   # red
    notable_down = PatternFill("solid", fgColor="FFFCE4D6")  # peach
    header_fill = PatternFill("solid", fgColor="FF1F3864")
    header_font = Font(bold=True, color="FFFFFFFF")
    bold = Font(bold=True)

    # rWAR only exists for pitchers, so a batters-only sheet keeps its old
    # shape and no reader is left wondering why two columns are empty.
    show_rwar = any(f.subject.rwar is not None for f in findings)
    columns = list(COLUMNS)
    if show_rwar:
        columns[7:7] = PITCHER_COLUMNS
    analysis_width = len(columns)
    columns += [(name, max(7, len(name) + 2)) for name in rating_columns]

    headers = [name for name, _width in columns]
    headers[5] = grade_label
    sheet.append(headers)
    for index, (_name, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width

    differential_column = headers.index("Differential") + 1
    numeric_columns = {headers.index(name) + 1 for name in
                       ("Projected WAR", "Expected WAR", "Differential", "z",
                        "rWAR", "rWAR - WAR") if name in headers}

    for rank, finding in enumerate(findings, start=1):
        subject = finding.subject
        age = subject.meta.get("age", "")
        values = [
            rank,
            subject.name,
            subject.position,
            int(age) if str(age).isdigit() else age,
            finding.group,
            subject.grade,
            round(subject.war, 2),
        ]
        if show_rwar:
            values += [
                None if subject.rwar is None else round(subject.rwar, 2),
                None if subject.war_gap is None else round(subject.war_gap, 2),
            ]
        values += [
            round(finding.expected_war, 2),
            round(finding.residual, 2),
            round(finding.z_score, 2),
            finding.scouting_accuracy,
        ]
        for column in rating_columns:
            raw = subject.ratings.get(column, "")
            try:
                values.append(float(raw) if raw not in ("", "-") else raw)
            except (TypeError, ValueError):
                values.append(raw)
        sheet.append(values)

        row = sheet.max_row
        z = finding.z_score
        if z >= strong_z:
            fill = strong_up
        elif z >= notable_z:
            fill = notable_up
        elif z <= -strong_z:
            fill = strong_down
        elif z <= -notable_z:
            fill = notable_down
        else:
            fill = None
        for column in range(1, len(columns) + 1):
            cell = sheet.cell(row=row, column=column)
            if fill is not None:
                cell.fill = fill
            if column in numeric_columns:
                cell.number_format = "0.00"
            if column == differential_column and abs(z) >= strong_z:
                cell.font = bold

    sheet.freeze_panes = "C2"
    if findings:
        sheet.auto_filter.ref = (f"A1:{get_column_letter(len(columns))}"
                                 f"{sheet.max_row}")
    return analysis_width


def write_xlsx(path: str, findings: list[Finding], fits: list[GroupFit],
               grade_label: str = "OVR", strong_z: float = STRONG_Z,
               notable_z: float = NOTABLE_Z,
               rating_columns: list[str] | None = None) -> None:
    openpyxl = _require_openpyxl()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Players"
    _write_players_sheet(sheet, findings, grade_label, strong_z, notable_z,
                         list(rating_columns or []))

    _write_method_sheet(workbook, fits, grade_label, strong_z, notable_z,
                        len(findings))
    workbook.save(path)


def _write_method_sheet(workbook, fits: list[GroupFit], grade_label: str,
                        strong_z: float, notable_z: float,
                        player_count: int) -> None:
    """A second sheet recording how the numbers were produced.

    Without it the Targets sheet is a list of assertions with no provenance,
    and the fitted line is the thing most worth arguing with.
    """
    from openpyxl.styles import Font

    sheet = workbook.create_sheet("How this was calculated")
    bold = Font(bold=True)
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 60

    def row(label, value=""):
        sheet.append([label, value])
        sheet.cell(row=sheet.max_row, column=1).font = bold

    row("What this measures",
        "How far a player's projected WAR sits above the WAR his grade "
        "predicts - not raw WAR.")
    row("Players sheet", "Every player in the pool, best differential first, "
                         "with his ratings alongside.")
    row("Green / amber", f"Underrated - projects above his grade. Green: z >= "
                         f"{strong_z}. Amber: z >= {notable_z}.")
    row("Red / peach", f"Overrated - projects below his grade. Red: z <= "
                       f"-{strong_z}. Peach: z <= -{notable_z}.")
    row("Uncoloured", "Within one standard deviation of the line - the grade "
                      "and the projection broadly agree.")
    row("Players listed", player_count)
    row("Caution", "Scouting accuracy is reported, never filtered on. A large "
                   "differential on a Low-accuracy report is a guess about a guess.")
    sheet.append([])

    for fit in fits:
        row(f"Fit: {fit.group}")
        sheet.append(["  Players", fit.count])
        sheet.append([f"  WAR per {grade_label} point", round(fit.slope, 4)])
        sheet.append(["  Intercept", round(fit.coefficients[0], 3)])
        sheet.append(["  Residual sd (1 z)", round(fit.residual_sd, 3)])
        if fit.note:
            sheet.append(["  Note", fit.note])
        offsets = fit.position_offsets
        if offsets:
            sheet.append(["  Position offsets",
                          f"wins vs {fit.reference_position} (the reference)"])
            for name, value in sorted(offsets.items(), key=lambda p: -p[1]):
                sheet.append([f"    {name}", round(value, 3)])
        elif fit.count:
            sheet.append(["  Position offsets",
                          "none - too few players per position to support them"])
        sheet.append([])
