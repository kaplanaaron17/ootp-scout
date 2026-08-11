"""The formatted .xlsx output."""

import os
import tempfile
import unittest

from ootp_scout import flagging, spreadsheet

try:
    import openpyxl
except ImportError:  # pragma: no cover - environment dependent
    openpyxl = None


_DEFAULT = object()


def finding(name, z, position="CF", grade=50.0, war=3.0, accuracy="High",
            ratings=None, implied=_DEFAULT):
    """`implied=None` means genuinely absent; omitting it derives a value."""
    subject = flagging.Subject(name=name, position=position, grade=grade,
                               war=war, meta={"scouting_accuracy": accuracy,
                                              "age": "22"},
                               ratings=ratings or {})
    implied_grade = grade + 5 * z if implied is _DEFAULT else implied
    return flagging.Finding(subject=subject, expected_war=war - z,
                            residual=z, z_score=z, group="hitters",
                            implied_grade=implied_grade)


def sample_fit():
    return flagging.GroupFit(group="hitters", count=40,
                             coefficients=[-10.0, 0.22, 0.8], residual_sd=1.9,
                             degree=1, positions=["C"], reference_position="1B")


GREEN, AMBER, RED, PEACH = "FFC7EFCE", "FFFFF2CC", "FFFFC7CE", "FFFCE4D6"


@unittest.skipIf(openpyxl is None, "openpyxl is not installed")
class WriteXlsxTest(unittest.TestCase):
    def setUp(self):
        self.findings = [
            finding("Strong Up", 3.1, ratings={"CON": "65", "POW": "70"}),
            finding("Notable Up", 1.4, ratings={"CON": "55", "POW": "50"}),
            finding("Ordinary", 0.2, ratings={"CON": "45", "POW": "45"}),
            finding("Notable Down", -1.4, ratings={"CON": "40", "POW": "35"}),
            finding("Strong Down", -2.6, ratings={"CON": "30", "POW": "25"}),
        ]
        self.folder = tempfile.TemporaryDirectory()
        self.path = os.path.join(self.folder.name, "pool.xlsx")
        spreadsheet.write_xlsx(self.path, self.findings, [sample_fit()],
                               grade_label="OVR",
                               rating_columns=["CON", "POW"])
        self.book = openpyxl.load_workbook(self.path)
        self.sheet = self.book["Players"]

    def tearDown(self):
        self.book.close()
        self.folder.cleanup()

    def _fill(self, row):
        return self.sheet.cell(row=row, column=2).fill.fgColor.rgb

    def test_one_sheet_holds_every_player(self):
        self.assertEqual(self.book.sheetnames,
                         ["Players", "How this was calculated"])
        self.assertEqual(self.sheet.max_row, 6)  # header plus five players

    def test_underrated_are_green_and_amber(self):
        self.assertEqual(self._fill(2), GREEN)
        self.assertEqual(self._fill(3), AMBER)

    def test_overrated_are_red_and_peach(self):
        self.assertEqual(self._fill(6), RED)
        self.assertEqual(self._fill(5), PEACH)

    def test_the_middle_is_left_uncoloured(self):
        self.assertNotIn(self._fill(4), (GREEN, AMBER, RED, PEACH))

    def test_ratings_are_included_and_numeric(self):
        headers = [cell.value for cell in self.sheet[1]]
        self.assertEqual(headers[-2:], ["CON", "POW"])
        self.assertEqual(self.sheet.cell(row=2, column=len(headers) - 1).value,
                         65)

    def test_grade_label_is_used(self):
        headers = [c.value for c in self.sheet[1]]
        self.assertEqual(headers[5], "OVR")
        self.assertEqual(headers[6], "Implied OVR")
        self.assertEqual(headers[7], "Implied - OVR")

    def test_implied_grade_and_gap_are_written(self):
        # Strong Up: grade 50, z 3.1, so implied 50 + 15.5 = 65.5 -> 66.
        self.assertEqual(self.sheet.cell(row=2, column=7).value, 66)
        self.assertEqual(self.sheet.cell(row=2, column=8).value, 16)

    def test_an_overrated_player_implies_a_lower_grade(self):
        self.assertLess(self.sheet.cell(row=6, column=8).value, 0)

    def test_a_missing_implied_grade_leaves_the_cells_empty(self):
        with tempfile.TemporaryDirectory() as folder:
            path = os.path.join(folder, "flat.xlsx")
            spreadsheet.write_xlsx(path, [finding("Flat", 0.5, implied=None)],
                                   [sample_fit()])
            book = openpyxl.load_workbook(path)
            try:
                sheet = book["Players"]
                self.assertIsNone(sheet.cell(row=2, column=7).value)
                self.assertIsNone(sheet.cell(row=2, column=8).value)
            finally:
                book.close()

    def test_rows_keep_the_order_they_were_given(self):
        names = [self.sheet.cell(row=r, column=2).value
                 for r in range(2, self.sheet.max_row + 1)]
        self.assertEqual(names, ["Strong Up", "Notable Up", "Ordinary",
                                 "Notable Down", "Strong Down"])

    def test_threshold_is_configurable(self):
        with tempfile.TemporaryDirectory() as folder:
            path = os.path.join(folder, "strict.xlsx")
            spreadsheet.write_xlsx(path, self.findings, [sample_fit()],
                                   strong_z=5.0, rating_columns=["CON"])
            book = openpyxl.load_workbook(path)
            try:
                sheet = book["Players"]
                fills = {sheet.cell(row=r, column=2).fill.fgColor.rgb
                         for r in range(2, sheet.max_row + 1)}
                self.assertNotIn(GREEN, fills)
                self.assertNotIn(RED, fills)
            finally:
                book.close()

    def test_method_sheet_explains_both_directions(self):
        sheet = self.book["How this was calculated"]
        text = "\n".join(str(cell.value) for row in sheet.iter_rows()
                         for cell in row if cell.value is not None)
        self.assertIn("Underrated", text)
        self.assertIn("Overrated", text)
        self.assertIn("0.22", text)      # slope
        self.assertIn("C", text)         # the position with an offset

    def test_freeze_panes_keeps_names_visible(self):
        self.assertEqual(self.sheet.freeze_panes, "C2")

    def test_empty_findings_still_produce_a_readable_file(self):
        with tempfile.TemporaryDirectory() as folder:
            path = os.path.join(folder, "empty.xlsx")
            spreadsheet.write_xlsx(path, [], [sample_fit()])
            book = openpyxl.load_workbook(path)
            try:
                self.assertEqual(book["Players"].max_row, 1)
            finally:
                book.close()

    def test_no_rating_columns_is_fine(self):
        with tempfile.TemporaryDirectory() as folder:
            path = os.path.join(folder, "bare.xlsx")
            spreadsheet.write_xlsx(path, self.findings, [sample_fit()])
            book = openpyxl.load_workbook(path)
            try:
                headers = [c.value for c in book["Players"][1]]
                self.assertEqual(headers[-1], "Scouting Accuracy")
            finally:
                book.close()


if __name__ == "__main__":
    unittest.main()
