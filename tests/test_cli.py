"""End-to-end tests over real ootpcalculator.com output.

`tests/fixtures/pool_batters.tsv` is a 41-player Batting Ratings export.
`tests/fixtures/batter-projections.csv` is what ootpcalculator.com actually
returned for it. The pool contains one planted player, Sleeper Sam, whose tools
are elite but whose OVR is 35 - the case the whole tool exists to catch.

The projections file also carries three players that are not in the report,
left over from a separate session, which is exactly the kind of mismatch a real
workflow produces.
"""

import io
import os
import tempfile
import unittest
from contextlib import contextmanager, redirect_stderr, redirect_stdout

from ootp_scout import clipboard, cli

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")
REPORT = os.path.join(FIXTURES, "pool_batters.tsv")
PROJECTIONS = os.path.join(FIXTURES, "batter-projections.csv")
PITCHER_REPORT = os.path.join(FIXTURES, "pool_pitchers.tsv")
PITCHER_PROJECTIONS = os.path.join(FIXTURES, "pitching-projections.csv")


# Commands that touch the database. Without an explicit --db they would use
# the real ootp_scout.db beside the tool, so a test run would write into the
# user's own league data - and, as it turned out, make the suite twenty times
# slower as that file grew. Every run is pointed at a throwaway file instead.
_DB_COMMANDS = {"flag", "report", "lookup", "stats", "leagues", "forget"}
_SCRATCH = tempfile.TemporaryDirectory()


def run(argv):
    if argv and argv[0] in _DB_COMMANDS and "--db" not in argv:
        argv = list(argv) + ["--db", os.path.join(_SCRATCH.name, "scratch.db")]
    out, err = io.StringIO(), io.StringIO()
    with redirect_stdout(out), redirect_stderr(err):
        code = cli.main(argv)
    return code, out.getvalue(), err.getvalue()


@contextmanager
def captured_clipboard(fail_with=None):
    """Intercept clipboard writes so tests never touch the real clipboard."""
    recorded = []
    original = cli.clipboard.copy

    def fake(text):
        if fail_with is not None:
            raise fail_with
        recorded.append(text)

    cli.clipboard.copy = fake
    try:
        yield recorded
    finally:
        cli.clipboard.copy = original


class FlagCommandTest(unittest.TestCase):
    def test_planted_sleeper_ranks_first(self):
        code, out, _ = run(["flag", REPORT, PROJECTIONS, "--limit", "5"])
        self.assertEqual(code, 0)
        body = out.splitlines()
        first_row = next(line for line in body if line.strip().startswith("1 "))
        self.assertIn("Sleeper Sam", first_row)

    def test_scouting_accuracy_is_shown_for_the_flagged_player(self):
        _, out, _ = run(["flag", REPORT, PROJECTIONS, "--limit", "5"])
        sleeper = next(line for line in out.splitlines() if "Sleeper Sam" in line)
        self.assertIn("Low", sleeper)

    def test_every_report_player_matches_a_projection(self):
        _, out, _ = run(["flag", REPORT, PROJECTIONS])
        self.assertIn("Matched: 41 of 41", out)

    def test_projection_only_players_do_not_appear(self):
        """Three players in the CSV are not in the report and must be ignored."""
        _, out, _ = run(["flag", REPORT, PROJECTIONS, "--limit", "50"])
        for absent in ("Cy Charlie", "Bo Bravo", "Al Alpha"):
            self.assertNotIn(absent, out)

    def test_min_z_filters(self):
        """--min-z applies to the underrated list, not the overrated one."""
        _, out, _ = run(["flag", REPORT, PROJECTIONS, "--min-z", "3",
                         "--overrated", "0"])
        rows = [line for line in out.splitlines()
                if line.strip() and line.strip()[0].isdigit()]
        self.assertEqual(len(rows), 1)
        self.assertIn("Sleeper Sam", rows[0])

    def test_writes_csv(self):
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "flagged.csv")
            code, _, _ = run(["flag", REPORT, PROJECTIONS, "--limit", "3",
                              "--out", destination])
            self.assertEqual(code, 0)
            with open(destination, encoding="utf-8") as handle:
                content = handle.read()
        self.assertIn("scouting_accuracy", content.splitlines()[0])
        self.assertIn("Sleeper Sam", content)

    def test_implied_grade_is_printed(self):
        _, out, _ = run(["flag", REPORT, PROJECTIONS, "--limit", "1",
                         "--overrated", "0"])
        self.assertIn("Impl", out)
        sleeper = next(l for l in out.splitlines() if "Sleeper Sam" in l)
        # OVR 35 against a 7.30 WAR projection implies a far higher grade.
        cells = sleeper.split()
        grade, implied = float(cells[-8]), float(cells[-7])
        self.assertEqual(grade, 35.0)
        self.assertGreater(implied, 60.0)

    def test_implied_grade_reaches_the_csv(self):
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "out.csv")
            run(["flag", REPORT, PROJECTIONS, "--limit", "3",
                 "--out", destination])
            with open(destination, encoding="utf-8") as handle:
                rows = handle.read().splitlines()
        self.assertIn("implied_grade", rows[0])
        self.assertIn("grade_gap", rows[0])
        columns = rows[0].split(",")
        first = rows[1].split(",")
        self.assertTrue(first[columns.index("implied_grade")])
        self.assertTrue(first[columns.index("grade_gap")].startswith("+"))

    def test_implied_grade_reaches_the_spreadsheet(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "pool.xlsx")
            run(["flag", REPORT, PROJECTIONS, "--out", destination])
            book = openpyxl.load_workbook(destination)
            try:
                sheet = book["Players"]
                headers = [c.value for c in sheet[1]]
                self.assertIn("Implied OVR", headers)
                column = headers.index("Implied OVR") + 1
                self.assertIsInstance(sheet.cell(row=2, column=column).value,
                                      (int, float))
            finally:
                book.close()

    def test_pitchers_get_rwar_columns(self):
        _, out, _ = run(["flag", PITCHER_REPORT, PITCHER_PROJECTIONS,
                         "--limit", "3", "--overrated", "0"])
        self.assertIn("rWAR", out)
        self.assertIn("R-W", out)

    def test_batters_do_not_get_rwar_columns(self):
        """Batter output has no IP, so the columns would only ever be empty."""
        _, out, _ = run(["flag", REPORT, PROJECTIONS, "--limit", "3",
                         "--overrated", "0"])
        self.assertNotIn("rWAR", out)

    def test_rwar_reaches_the_spreadsheet(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "arms.xlsx")
            run(["flag", PITCHER_REPORT, PITCHER_PROJECTIONS, "--limit", "5",
                 "--overrated", "0", "--out", destination])
            book = openpyxl.load_workbook(destination)
            try:
                sheet = book["Players"]
                headers = [c.value for c in sheet[1]]
                self.assertIn("rWAR", headers)
                self.assertIn("rWAR - WAR", headers)
                column = headers.index("rWAR") + 1
                self.assertIsInstance(sheet.cell(row=2, column=column).value,
                                      (int, float))
            finally:
                book.close()

    def test_rwar_reaches_the_csv(self):
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "arms.csv")
            run(["flag", PITCHER_REPORT, PITCHER_PROJECTIONS, "--limit", "3",
                 "--out", destination])
            with open(destination, encoding="utf-8") as handle:
                header, first = handle.read().splitlines()[:2]
        self.assertIn("rwar", header)
        self.assertIn("rwar_minus_war", header)
        self.assertTrue(first.split(",")[7])   # rwar column is populated

    def test_pitcher_pool_runs_end_to_end(self):
        code, out, _ = run(["flag", PITCHER_REPORT, PITCHER_PROJECTIONS,
                            "--limit", "5"])
        self.assertEqual(code, 0)
        self.assertIn("Matched: 31 of 31", out)
        self.assertIn("Ace Sleeper", out)

    def test_overrated_leaderboard_is_printed(self):
        _, out, _ = run(["flag", REPORT, PROJECTIONS, "--limit", "3",
                         "--overrated", "3"])
        self.assertIn("MOST UNDERRATED", out)
        self.assertIn("MOST OVERRATED", out)
        overrated_block = out.split("MOST OVERRATED")[1]
        rows = [line for line in overrated_block.splitlines()
                if line.strip() and line.strip()[0].isdigit()]
        self.assertEqual(len(rows), 3)
        # The worst shortfall leads, so the differential column is negative.
        self.assertIn("-", rows[0].split()[-3])

    def test_overrated_zero_turns_it_off(self):
        _, out, _ = run(["flag", REPORT, PROJECTIONS, "--overrated", "0"])
        self.assertNotIn("MOST OVERRATED", out)

    def test_spreadsheet_holds_the_whole_pool_regardless_of_limit(self):
        """--limit trims the printed tables; the sheet keeps everyone."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "pool.xlsx")
            _, out, _ = run(["flag", REPORT, PROJECTIONS, "--limit", "3",
                             "--out", destination])
            book = openpyxl.load_workbook(destination)
            try:
                self.assertEqual(book.sheetnames,
                                 ["Players", "How this was calculated"])
                self.assertEqual(book["Players"].max_row, 42)   # 41 players
            finally:
                book.close()
        self.assertIn("all 41 players", out)

    def test_spreadsheet_tints_both_directions(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "pool.xlsx")
            run(["flag", REPORT, PROJECTIONS, "--out", destination])
            book = openpyxl.load_workbook(destination)
            try:
                sheet = book["Players"]
                fills = [sheet.cell(row=r, column=2).fill.fgColor.rgb
                         for r in range(2, sheet.max_row + 1)]
            finally:
                book.close()
        self.assertIn("FFC7EFCE", fills)   # an underrated player, green
        self.assertIn("FFFFC7CE", fills)   # an overrated player, red

    def test_spreadsheet_carries_the_tool_ratings(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "pool.xlsx")
            run(["flag", REPORT, PROJECTIONS, "--out", destination])
            book = openpyxl.load_workbook(destination)
            try:
                headers = [c.value for c in book["Players"][1]]
            finally:
                book.close()
        for column in ("CON", "POW", "EYE", "DEF"):
            self.assertIn(column, headers)

    def test_xlsx_out_writes_a_real_spreadsheet(self):
        """--out foo.xlsx must produce a workbook, not CSV text renamed.

        An earlier version wrote CSV to the .xlsx path. Excel then refused to
        open it, and nothing caught it because the spreadsheet tests called
        write_xlsx directly and never went through the CLI.
        """
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "targets.xlsx")
            code, out, _ = run(["flag", REPORT, PROJECTIONS, "--limit", "5",
                                "--out", destination])
            self.assertEqual(code, 0)
            with open(destination, "rb") as handle:
                magic = handle.read(4)
        # Every .xlsx is a zip archive.
        self.assertEqual(magic, b"PK\x03\x04")
        self.assertIn("underrated", out)
        self.assertIn("overrated", out)

    def test_xlsx_out_is_readable_as_a_workbook(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "targets.xlsx")
            run(["flag", REPORT, PROJECTIONS, "--limit", "5",
                 "--out", destination])
            book = openpyxl.load_workbook(destination)
            try:
                self.assertIn("Players", book.sheetnames)
                self.assertIn("How this was calculated", book.sheetnames)
                sheet = book["Players"]
                self.assertEqual(sheet.max_row, 42)  # the whole pool
                self.assertEqual(sheet.cell(row=1, column=6).value, "OVR")
            finally:
                book.close()

    def test_csv_out_still_writes_csv(self):
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "targets.csv")
            run(["flag", REPORT, PROJECTIONS, "--limit", "3",
                 "--out", destination])
            with open(destination, encoding="utf-8") as handle:
                first = handle.readline()
        self.assertTrue(first.startswith("rank,name,team,position"))

    def test_mismatched_projections_file_fails_clearly(self):
        code, _, err = run(["flag", REPORT, REPORT])
        self.assertEqual(code, 1)
        self.assertIn("Download CSV", err)

    def test_projections_from_a_different_pool_are_refused(self):
        """Two unrelated exports must not produce a confident-looking ranking."""
        with tempfile.TemporaryDirectory() as folder:
            other = os.path.join(folder, "other-projections.csv")
            with open(PROJECTIONS, encoding="utf-8") as handle:
                lines = handle.read().splitlines()
            renamed = [lines[0]]
            for index, line in enumerate(lines[1:]):
                cells = line.split(",")
                cells[0] = f"Nobody {index}"
                renamed.append(",".join(cells))
            with open(other, "w", encoding="utf-8") as handle:
                handle.write("\n".join(renamed) + "\n")
            code, _, err = run(["flag", REPORT, other])
        self.assertEqual(code, 1)
        self.assertIn("two different pools", err)
        self.assertIn("41 players", err)

    def test_a_partial_match_warns_but_still_ranks(self):
        with tempfile.TemporaryDirectory() as folder:
            partial = os.path.join(folder, "partial-projections.csv")
            with open(PROJECTIONS, encoding="utf-8") as handle:
                lines = handle.read().splitlines()
            keep = [lines[0]]
            for index, line in enumerate(lines[1:]):
                cells = line.split(",")
                if index % 3:                      # drop two thirds of them
                    cells[0] = f"Nobody {index}"
                keep.append(",".join(cells))
            with open(partial, "w", encoding="utf-8") as handle:
                handle.write("\n".join(keep) + "\n")
            code, out, err = run(["flag", REPORT, partial, "--limit", "3"])
        self.assertEqual(code, 0)
        self.assertIn("WARNING", err)
        self.assertIn("matched a projection", err)
        self.assertTrue(any(line.strip().startswith("1 ")
                            for line in out.splitlines()))


class DatabaseCommandTest(unittest.TestCase):
    """flag records what it saw; lookup and report read it back."""

    def setUp(self):
        self.folder = tempfile.TemporaryDirectory()
        self.db = os.path.join(self.folder.name, "test.db")
        run(["flag", REPORT, PROJECTIONS, "--overrated", "0", "--db", self.db])
        run(["flag", PITCHER_REPORT, PITCHER_PROJECTIONS, "--overrated", "0",
             "--db", self.db])

    def tearDown(self):
        self.folder.cleanup()

    def test_flag_records_both_pools(self):
        _, out, _ = run(["stats", "--db", self.db])
        self.assertIn("72 players", out)
        self.assertIn("batter", out)
        self.assertIn("pitcher", out)

    def test_no_save_leaves_the_database_alone(self):
        with tempfile.TemporaryDirectory() as folder:
            db = os.path.join(folder, "untouched.db")
            run(["flag", REPORT, PROJECTIONS, "--no-save", "--db", db])
            _, out, _ = run(["stats", "--db", db])
        self.assertIn("empty", out)

    def test_lookup_finds_a_player(self):
        code, out, _ = run(["lookup", "Sleeper Sam", "--db", self.db])
        self.assertEqual(code, 0)
        self.assertIn("Sleeper Sam", out)
        self.assertIn("observation", out)

    def test_lookup_lists_candidates_for_a_partial_name(self):
        _, out, _ = run(["lookup", "sleeper", "--db", self.db])
        self.assertIn("Sleeper Sam", out)
        self.assertIn("Ace Sleeper", out)

    def test_lookup_on_an_unknown_player_explains(self):
        code, _, err = run(["lookup", "Nobody At All", "--db", self.db])
        self.assertEqual(code, 1)
        self.assertIn("Run `flag`", err)

    def test_report_refits_across_the_whole_database(self):
        code, out, _ = run(["report", "--limit", "5", "--db", self.db])
        self.assertEqual(code, 0)
        self.assertIn("72 players", out)
        # Both pools' planted players surface in one combined ranking.
        self.assertIn("Sleeper Sam", out)
        self.assertIn("Ace Sleeper", out)

    def test_report_keeps_roles_fitted_separately(self):
        _, out, _ = run(["report", "--db", self.db])
        self.assertIn("fit[hitters]", out)
        self.assertIn("fit[pitchers]", out)

    def test_report_can_restrict_to_one_role(self):
        _, out, _ = run(["report", "--role", "pitcher", "--db", self.db])
        self.assertIn("fit[pitchers]", out)
        self.assertNotIn("fit[hitters]", out)

    def test_report_writes_a_spreadsheet(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "db.xlsx")
            code, _, _ = run(["report", "--out", destination, "--db", self.db])
            self.assertEqual(code, 0)
            book = openpyxl.load_workbook(destination)
            try:
                self.assertEqual(book["Players"].max_row, 73)  # 72 players
            finally:
                book.close()

    def test_report_on_an_empty_database_explains(self):
        with tempfile.TemporaryDirectory() as folder:
            db = os.path.join(folder, "empty.db")
            code, _, err = run(["report", "--db", db])
        self.assertEqual(code, 1)
        self.assertIn("empty", err)

    def test_rerunning_the_same_pool_does_not_duplicate(self):
        run(["flag", REPORT, PROJECTIONS, "--overrated", "0", "--db", self.db])
        _, out, _ = run(["stats", "--db", self.db])
        self.assertIn("72 players", out)
        self.assertIn("72 observations", out)


class CompareCommandTest(unittest.TestCase):
    """Weighing two sides of a hypothetical trade."""

    def setUp(self):
        self.folder = tempfile.TemporaryDirectory()
        self.db = os.path.join(self.folder.name, "trade.db")
        run(["flag", REPORT, PROJECTIONS, "--overrated", "0",
             "--league", "Top Shelf", "--db", self.db])

    def tearDown(self):
        self.folder.cleanup()

    def _compare(self, a, b, *extra):
        return run(["compare", a, b, "--db", self.db, *extra])

    def test_both_sides_are_listed_with_totals(self):
        code, out, _ = self._compare("Sleeper Sam", "Player 02")
        self.assertEqual(code, 0)
        self.assertIn("YOU GIVE UP", out)
        self.assertIn("YOU RECEIVE", out)
        self.assertIn("Sleeper Sam", out)
        self.assertIn("Player 02", out)
        self.assertEqual(out.count("TOTAL"), 2)

    def test_multiple_players_a_side(self):
        _, out, _ = self._compare("Sleeper Sam, Player 15",
                                  "Player 02, Player 28")
        for name in ("Sleeper Sam", "Player 15", "Player 02", "Player 28"):
            self.assertIn(name, out)

    def test_the_side_with_more_projection_is_named(self):
        _, out, _ = self._compare("Player 15", "Sleeper Sam")
        self.assertIn("in your favour", out)

    def test_receiving_less_projection_is_named_too(self):
        _, out, _ = self._compare("Sleeper Sam", "Player 15")
        self.assertIn("against you", out)

    def test_the_grade_gap_is_reported_separately_from_war(self):
        """Winning on production and losing on value is the interesting case."""
        _, out, _ = self._compare("Sleeper Sam", "Player 02, Player 28")
        self.assertIn("Projected WAR", out)
        self.assertIn("Versus the grades", out)

    def test_it_says_plainly_that_this_is_not_surplus_value(self):
        _, out, _ = self._compare("Sleeper Sam", "Player 02")
        self.assertIn("not surplus value", out)

    def test_partial_names_resolve(self):
        code, out, _ = self._compare("sleeper", "Player 02")
        self.assertEqual(code, 0)
        self.assertIn("Sleeper Sam", out)

    def test_an_unknown_player_fails_with_a_pointer(self):
        code, _, err = self._compare("Nobody At All", "Player 02")
        self.assertEqual(code, 1)
        self.assertIn("Nobody At All", err)
        self.assertIn("lookup", err)

    def test_an_ambiguous_name_is_refused_rather_than_guessed(self):
        code, _, err = self._compare("Player 0", "Player 02")
        self.assertEqual(code, 1)
        self.assertIn("no single player", err)

    def test_a_one_sided_giveaway_is_allowed(self):
        code, out, _ = self._compare("Sleeper Sam", "")
        self.assertEqual(code, 0)
        self.assertIn("(nobody)", out)

    def test_naming_nobody_at_all_is_refused(self):
        code, _, err = self._compare("", "")
        self.assertEqual(code, 1)
        self.assertIn("at least one player", err)

    def test_it_will_not_span_leagues(self):
        run(["flag", PITCHER_REPORT, PITCHER_PROJECTIONS, "--overrated", "0",
             "--league", "Sim Nation", "--db", self.db])
        code, _, err = self._compare("Sleeper Sam", "Player 02")
        self.assertEqual(code, 1)
        self.assertIn("never combined", err)

    def test_a_player_from_another_league_is_not_found(self):
        run(["flag", PITCHER_REPORT, PITCHER_PROJECTIONS, "--overrated", "0",
             "--league", "Sim Nation", "--db", self.db])
        code, _, err = self._compare("Sleeper Sam", "Ace Sleeper",
                                     "--league", "Top Shelf")
        self.assertEqual(code, 1)
        self.assertIn("Ace Sleeper", err)


class MultiLeagueTest(unittest.TestCase):
    """Two leagues on different rating scales must never share a fit."""

    def setUp(self):
        self.folder = tempfile.TemporaryDirectory()
        self.db = os.path.join(self.folder.name, "two.db")
        run(["flag", REPORT, PROJECTIONS, "--overrated", "0",
             "--league", "Top Shelf", "--db", self.db])
        run(["flag", PITCHER_REPORT, PITCHER_PROJECTIONS, "--overrated", "0",
             "--league", "Sim Nation", "--db", self.db])

    def tearDown(self):
        self.folder.cleanup()

    def test_leagues_lists_both(self):
        code, out, _ = run(["leagues", "--db", self.db])
        self.assertEqual(code, 0)
        self.assertIn("Top Shelf", out)
        self.assertIn("Sim Nation", out)

    def test_leagues_shows_the_scale_of_each(self):
        _, out, _ = run(["leagues", "--db", self.db])
        self.assertIn("20 to 80", out)

    def test_report_refuses_to_guess_between_leagues(self):
        code, _, err = run(["report", "--db", self.db])
        self.assertEqual(code, 1)
        self.assertIn("different rating scales", err)
        self.assertIn("Top Shelf", err)
        self.assertIn("Sim Nation", err)

    def test_report_works_when_told_which_league(self):
        code, out, _ = run(["report", "--league", "Top Shelf", "--limit", "3",
                            "--db", self.db])
        self.assertEqual(code, 0)
        self.assertIn("Top Shelf", out)
        self.assertIn("Sleeper Sam", out)
        # The other league's players must not leak into this fit.
        self.assertNotIn("Ace Sleeper", out)

    def test_a_partial_league_name_resolves(self):
        code, out, _ = run(["report", "--league", "sim", "--limit", "3",
                            "--db", self.db])
        self.assertEqual(code, 0)
        self.assertIn("Ace Sleeper", out)

    def test_an_unknown_league_lists_what_is_held(self):
        code, _, err = run(["report", "--league", "Nowhere", "--db", self.db])
        self.assertEqual(code, 1)
        self.assertIn("No league named", err)
        self.assertIn("Top Shelf", err)

    def test_stats_breaks_down_by_league(self):
        _, out, _ = run(["stats", "--db", self.db])
        self.assertIn("Top Shelf", out)
        self.assertIn("Sim Nation", out)

    def test_forget_removes_one_league_only(self):
        code, out, _ = run(["forget", "Top Shelf", "--yes", "--db", self.db])
        self.assertEqual(code, 0)
        self.assertIn("Deleted", out)
        _, listed, _ = run(["leagues", "--db", self.db])
        self.assertNotIn("Top Shelf", listed)
        self.assertIn("Sim Nation", listed)

    def test_forget_refuses_an_unknown_league(self):
        code, _, err = run(["forget", "Nowhere", "--yes", "--db", self.db])
        self.assertEqual(code, 1)
        self.assertIn("No league named", err)

    def test_after_forgetting_one_report_needs_no_league(self):
        run(["forget", "Top Shelf", "--yes", "--db", self.db])
        code, out, _ = run(["report", "--limit", "2", "--db", self.db])
        self.assertEqual(code, 0)
        self.assertIn("Sim Nation", out)

    def test_a_single_league_database_needs_no_flag(self):
        with tempfile.TemporaryDirectory() as folder:
            db = os.path.join(folder, "one.db")
            run(["flag", REPORT, PROJECTIONS, "--overrated", "0",
                 "--league", "Only One", "--db", db])
            code, out, _ = run(["report", "--limit", "2", "--db", db])
        self.assertEqual(code, 0)
        self.assertIn("Only One", out)


class PrepareCommandTest(unittest.TestCase):
    def test_detects_the_view_and_writes_a_paste_block(self):
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "paste.tsv")
            with captured_clipboard():
                code, out, _ = run(["prepare", REPORT, "--out", destination])
            self.assertEqual(code, 0)
            with open(destination, encoding="utf-8") as handle:
                lines = handle.read().splitlines()
        self.assertIn("Batting Ratings", out)
        self.assertIn("batting-projections.csv", out)
        self.assertEqual(len(lines), 42)  # header plus 41 players
        self.assertTrue(lines[0].startswith("POS\t#\tName"))

    def test_paste_block_goes_to_the_clipboard_by_default(self):
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "paste.tsv")
            with captured_clipboard() as recorded:
                _, out, _ = run(["prepare", REPORT, "--out", destination])
        self.assertEqual(len(recorded), 1)
        self.assertIn("Ctrl+V", out)
        with open(REPORT, encoding="utf-8") as handle:
            expected_players = len(handle.read().splitlines()) - 1
        self.assertEqual(len(recorded[0].strip().splitlines()),
                         expected_players + 1)
        self.assertTrue(recorded[0].startswith("POS\t#\tName"))

    def test_no_copy_leaves_the_clipboard_alone(self):
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "paste.tsv")
            with captured_clipboard() as recorded:
                _, out, _ = run(["prepare", REPORT, "--out", destination,
                                 "--no-copy"])
        self.assertEqual(recorded, [])
        self.assertNotIn("Ctrl+V", out)
        self.assertIn("Paste the whole contents", out)

    def test_a_clipboard_failure_is_reported_but_not_fatal(self):
        """A locked clipboard must not lose the run - the file is still there."""
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "paste.tsv")
            failure = clipboard.ClipboardError("clipboard busy")
            with captured_clipboard(fail_with=failure):
                code, out, err = run(["prepare", REPORT, "--out", destination])
            self.assertTrue(os.path.exists(destination))
        self.assertEqual(code, 0)
        self.assertIn("clipboard busy", err)
        self.assertIn("Paste the whole contents", out)

    def test_rejects_values_the_calculator_would_reject(self):
        with tempfile.TemporaryDirectory() as folder:
            bad = os.path.join(folder, "bad.tsv")
            with open(REPORT, encoding="utf-8") as handle:
                lines = handle.read().splitlines()
            cells = lines[1].split("\t")
            cells[8] = "52"  # CON, not a multiple of 5
            lines[1] = "\t".join(cells)
            with open(bad, "w", encoding="utf-8") as handle:
                handle.write("\n".join(lines) + "\n")
            code, _, err = run(["prepare", bad, "--scale", "20 to 80"])
        self.assertEqual(code, 1)
        self.assertIn("multiple of 5", err)
        self.assertIn('--scale "1 to 100"', err)

    def test_same_values_pass_on_the_1_100_scale(self):
        with tempfile.TemporaryDirectory() as folder:
            destination = os.path.join(folder, "paste.tsv")
            with captured_clipboard():
                code, _, _ = run(["prepare", REPORT, "--scale", "1 to 100",
                                  "--out", destination])
        self.assertEqual(code, 0)

    def test_draft_pool_export_is_rejected_with_guidance(self):
        with tempfile.TemporaryDirectory() as folder:
            path = os.path.join(folder, "draft.csv")
            with open(path, "w", encoding="utf-8") as handle:
                handle.write("POS,#,Name,Inf,DOB,Age,NAT,HT,WT,B,T,OVR,POT,"
                             "Prone,DEM,Sign,SctAcc\n")
                handle.write("SP,14,Nicolas Arnaud,,02/20/2010,21,CAN,6' 2\","
                             "185 lbs,Right,Right,40,80,Normal,-,Easy,High\n")
            code, _, err = run(["prepare", path])
        self.assertEqual(code, 1)
        self.assertIn("Ratings view", err)



class CrossRoleTest(unittest.TestCase):
    """Pitchers turn up in batting exports; their grade is not a hitting grade."""

    def _mixed(self):
        return [cli.flagging.Subject(name="Hitter", position="CF", grade=50,
                                     war=2.0),
                cli.flagging.Subject(name="Catcher", position="C", grade=50,
                                     war=2.0),
                cli.flagging.Subject(name="Starter", position="SP", grade=50,
                                     war=-4.0),
                cli.flagging.Subject(name="Reliever", position="RP", grade=50,
                                     war=-4.0)]

    def test_pitchers_are_dropped_from_a_batter_pool(self):
        out = io.StringIO()
        with redirect_stdout(out):
            kept = cli.drop_cross_role(self._mixed(), "batter")
        self.assertEqual([s.name for s in kept], ["Hitter", "Catcher"])
        self.assertIn("Ignoring 2 pitchers", out.getvalue())

    def test_position_players_are_dropped_from_a_pitcher_pool(self):
        with redirect_stdout(io.StringIO()):
            kept = cli.drop_cross_role(self._mixed(), "pitcher")
        self.assertEqual([s.name for s in kept], ["Starter", "Reliever"])

    def test_closers_count_as_pitchers(self):
        pool = [cli.flagging.Subject(name="Closer", position="CL", grade=50,
                                     war=1.0)]
        with redirect_stdout(io.StringIO()):
            self.assertEqual(len(cli.drop_cross_role(pool, "pitcher")), 1)
            self.assertEqual(len(cli.drop_cross_role(pool, "batter")), 0)

    def test_a_clean_pool_says_nothing(self):
        pool = [cli.flagging.Subject(name="Hitter", position="CF", grade=50,
                                     war=2.0)]
        out = io.StringIO()
        with redirect_stdout(out):
            cli.drop_cross_role(pool, "batter")
        self.assertEqual(out.getvalue(), "")


class ScaleBoundsTest(unittest.TestCase):
    def test_known_scales_map_to_their_limits(self):
        self.assertEqual(cli.bounds_for("20 to 80"), (20.0, 80.0))
        self.assertEqual(cli.bounds_for("1 to 100"), (1.0, 100.0))

    def test_an_unknown_or_missing_scale_gives_no_bounds(self):
        self.assertIsNone(cli.bounds_for(None))
        self.assertIsNone(cli.bounds_for(""))
        self.assertIsNone(cli.bounds_for("something else"))


class GradeFloorTest(unittest.TestCase):
    """Unusable players distort the fit; --min-grade excludes them."""

    def _subjects(self, grades_and_wars):
        return [cli.flagging.Subject(name=f"P{i}", position="CF",
                                     grade=g, war=w)
                for i, (g, w) in enumerate(grades_and_wars)]

    def test_the_floor_drops_players_below_it(self):
        subjects = self._subjects([(20, -8.0), (30, -3.0), (50, 2.0),
                                   (60, 4.0)])
        with redirect_stdout(io.StringIO()):
            kept = cli.apply_grade_floor(subjects, 40)
        self.assertEqual([s.grade for s in kept], [50, 60])

    def test_the_floor_says_how_many_it_dropped(self):
        subjects = self._subjects([(20, -8.0), (50, 2.0), (60, 4.0)])
        out = io.StringIO()
        with redirect_stdout(out):
            cli.apply_grade_floor(subjects, 40)
        self.assertIn("Ignoring 1 player", out.getvalue())

    def test_no_floor_keeps_everyone(self):
        subjects = self._subjects([(20, -8.0), (50, 2.0), (60, 4.0)])
        err = io.StringIO()
        with redirect_stderr(err):
            kept = cli.apply_grade_floor(subjects, None)
        self.assertEqual(len(kept), 3)

    def test_a_pool_full_of_sub_replacement_players_warns(self):
        subjects = self._subjects([(20, -8.0)] * 8 + [(60, 4.0)] * 2)
        err = io.StringIO()
        with redirect_stderr(err):
            cli.apply_grade_floor(subjects, None)
        self.assertIn("below replacement", err.getvalue())
        self.assertIn("--min-grade", err.getvalue())

    def test_a_healthy_pool_does_not_warn(self):
        subjects = self._subjects([(45, 1.0), (50, 2.0), (55, 3.0),
                                   (60, 4.0), (40, -0.5)])
        err = io.StringIO()
        with redirect_stderr(err):
            cli.apply_grade_floor(subjects, None)
        self.assertEqual(err.getvalue(), "")

    def test_the_floor_steepens_nothing_it_should_not(self):
        """The point of the floor: a low-grade mass flattens implied grades.

        Fitting only the usable players gives a shallower slope, which spreads
        the implied grades back out at the top instead of compressing them.
        """
        pool = [(20, -8.0)] * 40 + [(g, (g - 40) * 0.15) for g in
                                    range(40, 71, 5)]
        subjects = self._subjects(pool)
        flooded = cli.flagging.analyze(subjects, position_adjust=False)
        with redirect_stdout(io.StringIO()):
            kept = cli.apply_grade_floor(subjects, 40)
        clean = cli.flagging.analyze(kept, position_adjust=False)
        self.assertGreater(flooded.fits[0].slope, clean.fits[0].slope)
        self.assertGreater(clean.fits[0].implied_grade(4.0, "CF"),
                           flooded.fits[0].implied_grade(4.0, "CF"))

    def test_report_accepts_the_flag(self):
        with tempfile.TemporaryDirectory() as folder:
            db = os.path.join(folder, "floor.db")
            run(["flag", REPORT, PROJECTIONS, "--overrated", "0",
                 "--league", "L", "--db", db])
            code, out, _ = run(["report", "--min-grade", "45", "--limit", "3",
                                "--db", db])
        self.assertEqual(code, 0)
        self.assertIn("Ignoring", out)

    def test_a_floor_that_removes_everyone_fails_clearly(self):
        with tempfile.TemporaryDirectory() as folder:
            db = os.path.join(folder, "floor.db")
            run(["flag", REPORT, PROJECTIONS, "--overrated", "0",
                 "--league", "L", "--db", db])
            code, _, err = run(["report", "--min-grade", "99", "--db", db])
        self.assertEqual(code, 1)
        self.assertIn("grade floor", err)


if __name__ == "__main__":
    unittest.main()
